'''
Name: Automated ProviderOne Check
Author: Wesley Wang
Date: 10/28/2018
Description: Automate ProviderOne eligibility check with Selenium WebDriver
             using extracted client info from a Excel worksheet.
'''

import time
import os
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from xlrd import XLRDError
import pandas as pd


URL = "https://www.waproviderone.org/"  # ProviderOne URL
WB_PATH = "Clients.xlsx"   # Workbook file path
WS_NAME = "main"    # Target worksheet name
CLT_PATH = "."  # Client folder root path
ALT_PATH = "./Not Found"    # Backup export path if client folder is missing
ZOOM_LEVEL = "70%"


def create_folder(path):
    '''
    Create alternative folder to hold screenshots
    that do not have folder in destined location
    '''
    if not os.path.isdir(path):
        try:
            os.mkdir(path)
            full_path = os.path.abspath(path)
            print(f"Folder created at {full_path}")
        except OSError:
            print("An error occured when creating folder!")


def get_login():
    '''
    Obtain user's ProviderOne login information
    '''
    check_login = ""
    while check_login != "y":
        domain_id = input("Please enter your ProviderOne Domain ID: ")
        user_name = input("Please enter your ProviderOne Username: ")
        password = input("Please enter your ProviderOne Password: ")
        print(f'''Your login info is:
            Domain ID: {domain_id}
            Username: {user_name} 
            Password: {password}''')
        check_login = input("Is it correct? (Y/N) ").lower()
        if check_login not in ["y", "n"]:
            print("Please enter correct answer (Y/N)!!")
    login = (domain_id, user_name, password)
    return login


def page_load(browser, key):
    '''
    Wait for page to load until key element is clickable
    '''
    try:
        WebDriverWait(browser, 60).until(ec.element_to_be_clickable((By.ID, key)))
    except TimeoutException:
        print("Timeout")


def load_clients(file):
    '''
    Read clients information from target Excel worksheet
    Return a list of client info and a tuple that holds
    index of Prov1 Status and Prov1 Location columns
    '''
    try:
        sheet = pd.read_excel(file, WS_NAME, na_filter=False)
        clients = [(lname, fname, dob.strftime("%m/%d/%Y"), ssn.replace('-', '').strip())
                   for (lname, fname, dob, ssn)
                   in zip(sheet["LastName"], sheet["FirstName"], sheet["DoB"], sheet["SSN"], )]
        prov1_cols = (sheet.columns.get_loc("ProviderOne Status") + 1,
                      sheet.columns.get_loc("ProviderOne Location") + 1)
        return clients, prov1_cols
    except FileNotFoundError:
        print("File not found !")
    except XLRDError:
        print("An error occured when reading Excel file!")


def search_client(browser, sheet, date, row, stat_col,
                  loc_col, lname, fname, dob, ssn, use_ssn=False):
    '''
    Search client eligibility with given info
    Update worksheet to show eligibility status and location
    Take screenshot of eligbility status page
    then store it in client folder
    Remove previous screenshot of bad search result if applicable
    '''
    page_load(browser, "rfld_d:TO_DATE")

    if use_ssn:
        browser.find_element_by_id("nfld_n:dbZzMBR_IDENTIFIER-IDNTFR_SSN").send_keys(ssn)
    else:
        browser.find_element_by_id("nfld_a:dbZzMBR_DMGRPHC-LAST_NAME").send_keys(lname)
        browser.find_element_by_id("nfld_a:dbZzMBR_DMGRPHC-FIRST_NAME").send_keys(fname)
    browser.find_element_by_id("nfld_d:dbZzMBR_DMGRPHC-BIRTH_DATE").send_keys(dob)
    browser.execute_script("submitForm()")

    page_load(browser, "nlbl:CountyCode")

    if "Active Coverage" in browser.find_element_by_id("nlbl:BenifitInformationCode").text:
        file_name = "Inc-Res-Ins"
        sheet.cell(row=row, column=stat_col).value = "Prov1"
        sheet.cell(row=row, column=loc_col).value = \
            browser.find_element_by_id("nlbl:CountyCode").text.split("-")[1]
    elif "Inactive" in browser.find_element_by_id("nlbl:BenifitInformationCode").text:
        file_name = "Inactive_ProviderOne"
        sheet.cell(row=row, column=stat_col).value = "Inactive"
        sheet.cell(row=row, column=loc_col).value = "None"
    else:
        file_name = "Inactive_ProviderOne"
        if sheet.cell(row=row, column=stat_col).value is None:
            sheet.cell(row=row, column=stat_col).value = "None"
            sheet.cell(row=row, column=loc_col).value = "None"

    efile_path = CLT_PATH + f"/{lname}, {fname} {dob.replace('/', '-')}/Client Eligibility"
    if not os.path.isdir(efile_path):
        efile_path = ALT_PATH
        file_name = f"({fname} {lname})" + file_name
    file_ext = f"_{fname[0]}{lname[0]}_{date[0]}_{date[1]}.png"

    browser.execute_script(f"document.body.style.zoom = '{ZOOM_LEVEL}'")

    export_path = efile_path + "/" + file_name + file_ext
    if not os.path.exists(export_path):
        for doc in os.listdir(efile_path):
            if "Inactive" in doc and doc.endswith(file_ext):
                os.remove(os.path.join(efile_path, doc))
        browser.get_screenshot_as_file(export_path)

    browser.execute_script("submitForm('InquiryButton')")


def browse_prov1(login, clients, prov1_col):
    '''
    Log into ProviderOne
    Pass in client info to capture eligibility status
    '''
    browser = webdriver.Chrome("chromedriver.exe")
    browser.maximize_window()
    browser.get(URL)
    browser.find_element_by_id("rfld_a:DomainNme").send_keys(login[0])
    browser.find_element_by_id("rfld_a:UserID").send_keys(login[1])
    browser.find_element_by_id("nfld:UserPwd").send_keys(login[2])
    browser.find_element_by_id("LoginButton").click()
    browser.find_element_by_id("GoButton").click()
    browser.find_element_by_link_text("Benefit Inquiry").click()

    workbook = load_workbook(WB_PATH)
    sheet = workbook[WS_NAME]
    date = (time.strftime("%Y"), time.strftime("%m"), time.strftime("%d"))
    stat_col = prov1_col[0]
    loc_col = prov1_col[1]
    for client in clients:
        row = clients.index(client) + 2
        search_client(browser, sheet, date, row, stat_col, loc_col,
                      client[0], client[1], client[2], client[3])
        if len(client[3]) == 9 and client[3].isdigit() and \
            sheet.cell(row=row, column=stat_col).value == "None":
            search_client(browser, sheet, date, row, stat_col, loc_col,
                          client[0], client[1], client[2], client[3], True)

    workbook.save(WB_PATH)
    browser.close()


if __name__ == "__main__":
    create_folder(ALT_PATH)
    browse_prov1(get_login(), *load_clients(WB_PATH))
